import os.path
import re
import openpyxl


class Method:
    """方法类"""
    def __init__(self):
        """初始化"""
        # 初始化读取敏感表和库
        self.sensitive_table_dict = {}   # 库名:表名
        self.sensitive_table_dict_init()
        # 正则规则
        self.re_schema = re.compile("数据库：(.*?)，")
        self.re_table = re.compile("from(.*) ")
        # 刷新输出文件
        open("output.txt", "w").write("")

    def sensitive_table_dict_init(self):
        """敏感表"""
        path = "config/jioaoben.xlsx"
        book = openpyxl.load_workbook(path)
        sheet = book.worksheets[0]
        line = 0
        while True:
            line += 1
            order = sheet.cell(line, 1).value
            if line > 10 and order is None:
                break
            schema = sheet.cell(line, 3).value
            if schema is not None:
                schema = schema.lower()
            table = sheet.cell(line, 4).value
            if table is not None:
                table = table.lower()
            if schema in self.sensitive_table_dict:
                self.sensitive_table_dict[schema].append(table)
            else:
                self.sensitive_table_dict[schema] = [table]

    def output(self, message):
        """输出文件"""
        with open("output.txt", "a", encoding="utf-8") as f:
            f.write(f"{message}\n")

    def check_sql(self, line, sql):
        """检测SQL语句"""
        sql = sql.lower()
        schemas = self.re_schema.findall(sql)
        if len(schemas) == 0:
            return
        schema = schemas[0].strip()
        tables = sql.split(" ")
        for table in tables:
            table = table.strip()
            if "." in table:
                # print(table)
                tmp = table.split(".")
                if len(tmp) != 2:
                    # print(table)
                    continue
                tmp_schema, tmp_table = tmp
                tmp_schema = tmp_schema
                tmp_table = tmp_table
                if tmp_schema in self.sensitive_table_dict and tmp_table in self.sensitive_table_dict[tmp_schema]:
                    message = f"行数:{line}存在违规敏感数据。{tmp_schema}:{tmp_table}"
                    print(message)
                    self.output(message)
            else:
                if schema in self.sensitive_table_dict and table in self.sensitive_table_dict[schema]:
                    message = f"行数:{line}存在违规敏感数据。{schema}:{table}"
                    print(message)
                    self.output(message)
        tables = self.re_table.findall(sql)

        # for tmp_table in tables:
        #     if "where" in tmp_table:
        #         tables = tmp_table.split("where")
        #     elif ';' in tmp_table:
        #         tables = tmp_table.split(";")
        #     else:
        #         tables = [tmp_table]
        #     for table in tables:
        #         table = table.strip()
        #         if " " in table:
        #             table = table.split(" ")[0]
        #         if "." in table:
        #             tmp_schema, tmp_table = table.split(".")
        #             tmp_schema = tmp_schema
        #             tmp_table = tmp_table
        #             if tmp_schema in self.sensitive_table_dict and tmp_table in self.sensitive_table_dict[tmp_schema]:
        #                 message = f"行数:{line}存在违规敏感数据。{tmp_schema}:{tmp_table}"
        #                 print(message)
        #                 self.output(message)
        #         else:
        #             if schema in self.sensitive_table_dict and table in self.sensitive_table_dict[schema]:
        #                 message = f"行数:{line}存在违规敏感数据。{schema}:{table}"
        #                 print(message)
        #                 self.output(message)

    def check_xlsx(self, check_path):
        """检测表格"""
        try:
            book = openpyxl.load_workbook(check_path)
        except:
            print("文件损坏")
            return False
        sheet = book.worksheets[0]
        line = 1
        while True:
            line += 1
            order = sheet.cell(line, 1).value
            if order is None:
                break
            for num in range(1, 100):
                sql = sheet.cell(line, num).value
                self.check_sql(line, f"{sql}")

    def check_path(self, path):
        """检测路径文件，可能带双引号"""
        path = path.replace('"', "")
        ext = path.split(".")[-1]
        if ext != "xlsx":
            print("文件不是xlsx文件")
            return ""
        if not os.path.exists(path):
            print("文件不存在")
            return ""
        return path

    def test(self):
        """测试"""
        sql = "用户yx_dingmh1: 执行sql语句，数据源：dwi_bss，数据库：dwi_bss，SQL：select a.USER_ID,b.SIGN_SCENE_CODEfrom ptmp_yx.T_P_SD_CESHI_TEST3_12800_20230301 a left outer join (select * from dwi_bss.TW_D_CM_DM_M2M_CT_SUBS_BUSI where STATIS_YMD = 20230301 and PROV_ID = 12800 )b on a.USER_ID=b.USER_IDwhere a.IS_SIGN_SCENE =0 limit 10;"
        self.check_sql(1, sql)
        exit()

    def run(self):
        """运行类"""
        # self.test()
        while True:
            path = input("请输入要检测的xlsx表格的路径>")
            # check_path = "脚本.xlsx"
            new_path = self.check_path(path)
            if new_path == "":
                continue
            self.check_xlsx(new_path)



def main():
    """主程序"""
    method = Method()
    method.run()


if __name__ == '__main__':
    main()

    def check_xlsx(self, check_path):
        """检测表格"""
        try:
            book = openpyxl.load_workbook(check_path)
        except:
            print("文件损坏")
            return False
        sheet = book.worksheets[0]
        line = 1
        sensitive_rows = []
        while True:
            line += 1
            order = sheet.cell(line, 1).value
            if order is None:
                break
            for num in range(1, 100):
                sql = sheet.cell(line, num).value
                if self.check_sql(line, f"{sql}"):
                    sensitive_rows.append(line)
                    break

        # 创建新的工作簿和工作表
        new_book = openpyxl.Workbook()
        new_sheet = new_book.active

        # 将敏感行复制到新工作表中
        for row_index, row in enumerate(sensitive_rows):
            for col_index in range(1, sheet.max_column + 1):
                new_sheet.cell(row=row_index + 1, column=col_index).value = sheet.cell(row=row, column=col_index).value

        # 保存新工作簿
        new_book.save('sensitive_rows.xlsx')
